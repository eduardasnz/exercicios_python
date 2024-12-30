import openpyxl

# Criar uma planilha
book = openpyxl.Workbook()

# Visualizar páginas existentes
print(book.sheetnames)

# Criar página
book.create_sheet('Frutas')

# Selecionando uma página
frutas_page = book['Frutas']
frutas_page.append(['Frutas','Quantidade','Preço'])
frutas_page.append(['Maça','3','R$10,92'])
frutas_page.append(['Banana','7','R$5,92'])
frutas_page.append(['Kiwi','5','R$7,20'])
frutas_page.append(['Caju','2','R$1,9'])
frutas_page.append(['Manga','13','R$11,07'])

# Salvar planilha
book.save('Planilha de compras.xlsx')